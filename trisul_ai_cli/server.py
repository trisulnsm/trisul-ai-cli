import warnings
warnings.filterwarnings("ignore", category=FutureWarning, module="google")

from mcp.server.fastmcp import FastMCP
from trisul_ai_cli import trp_pb2
import zmq
import datetime
import sqlite3
import uuid
import chromadb
from google.protobuf.json_format import MessageToDict
import logging
from pathlib import Path
from datetime import datetime, timedelta, timezone
import random
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet
import os
import ast
from trisul_ai_cli.tools.json_to_toon_converter import json_to_toon
import json
from typing import List, Any
from dotenv import dotenv_values
from pathlib import Path
from trisul_ai_cli.llm_factory import LLMFactory


logging.basicConfig(
    filename= Path(os.getcwd()) / "trisul_ai_cli.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    force=True
)


global _global_zmq_context
_global_zmq_context = zmq.Context()


mcp = FastMCP(name="trisul-mcp-server")

FLOWINTFS_GUID = "{C0B04CA7-95FA-44EF-8475-3835F3314761}"
FLOWGENS_GUID = "{2314BB8E-2BCC-4B86-8AA2-677E5554C0FE}"

# Helper functions

def normalize_context(ctx: str) -> str:
    try:
        ctx = ctx.lower()
        logging.info(f"[normalize_context] Normalizing context: {ctx}")
        if ctx.startswith("context_"):
            ctx = ctx.split("_", 1)[-1]
        if ctx == "default" or ctx == "context0":
            normalized = "context0"
        else:
            normalized = f"context_{ctx}"
        logging.info(f"[normalize_context] Normalized context: {normalized}")
        return normalized
    except Exception as e:
        logging.error(f"[normalize_context] Error normalizing context '{ctx}': {str(e)}")
        return "context0"  # Default fallback


def countergroup_info(zmq_endpoint: str = None, context: str = "context0", get_meter_info: bool = False):
    """Fetch all counter groups information from Trisul via ZMQ for a given zmq_endpoint.
    and it will also fetch meters info for each counter group so that we can determine what each meter means and its index.
    for example if we want to get the counter group guid for "FlowIntfs" and meter index for "Received traffic" we can use this function.
    Example output:
        [{
            "guid": "{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}","name": "FlowIntfs","bucketSize": "60000","topperBucketSize": "300",
            "timeInterval": { "from": {"tvSec": "1718711400","tvUsec": "0"}, "to": {"tvSec": "1718712060","tvUsec": "0"} },
            "meters": [
                { "id": 0, "type": "VT_RATE_COUNTER", "topcount": 1000, "name": "Bps", "description": "Total", "units": "Bps" },
                { "id": 1, "type": "VT_RATE_COUNTER", "topcount": 1000, "name": "Bps", "description": "Recv", "units": "Bps" },
                { "id": 2, "type": "VT_RATE_COUNTER", "topcount": 1000, "name": "Bps", "description": "Xmit", "units": "Bps" }
            ]
        }, ...]
    
    Arguments: zmq_endpoint (str): ZMQ Endpoint
    Returns: dict: Dictionary containing counter group information.
    """
    try:
        logging.info(f"[countergroup_info] Connecting to zmq_endpoint: {zmq_endpoint}, get_meter_info: {get_meter_info}")        
        
        # to retrieve all counter groups send an empty COUNTER_GROUP_INFO_REQUEST
        try:
            logging.info("[countergroup_info] Building COUNTER_GROUP_INFO_REQUEST")
            req = trp_pb2.Message()
            req.trp_command = req.COUNTER_GROUP_INFO_REQUEST
            req.counter_group_info_request.get_meter_info = get_meter_info
        except Exception as e:
            logging.error(f"[countergroup_info] Error building request: {str(e)}")
            raise
        
        logging.info("[countergroup_info] Sending COUNTER_GROUP_INFO_REQUEST...")
        resp = get_response(zmq_endpoint, req)
        
        result = MessageToDict(resp)
        logging.info(f"[countergroup_info] Received response with {len(result.get('groupDetails', []))} groups")
        return result
        
    except Exception as e:
        logging.error(f"[countergroup_info] Error in countergroup_info: {str(e)}", exc_info=True)
        return {"error": str(e), "groupDetails": []}


def epoch_to_duration(from_ts, to_ts):
    
    secs = int(to_ts) - int(from_ts)
    
    lookup = {"days": "Day", "hours": "Hr", "minutes": "Min", "seconds": "Sec"}
    
    secs = int(secs)
    if secs == 0:
        return "<1s"
    
    duration = ""
    
    # Calculate days
    days = secs // 86400
    if days > 0:
        lookup["days"] = f" {lookup['days']}s " if days > 1 else f" {lookup['days']} "
        duration += f"{days}{lookup['days']}"
        secs = secs - days * 86400
    
    # Calculate hours
    hours = secs // 3600
    if hours > 0:
        lookup["hours"] = f" {lookup['hours']}s " if hours > 1 else f" {lookup['hours']} "
        duration += f"{hours}{lookup['hours']}"
        secs = secs - hours * 3600
    
    # Calculate minutes
    minutes = secs // 60
    if minutes > 0:
        lookup["minutes"] = f" {lookup['minutes']}s " if minutes > 1 else f" {lookup['minutes']} "
        duration += f"{minutes}{lookup['minutes']}"
        secs = secs - minutes * 60
    
    # Remaining seconds
    if secs > 0:
        lookup["seconds"] = f" {lookup['seconds']}s " if secs > 1 else f" {lookup['seconds']} "
        duration += f"{secs}{lookup['seconds']}"
    
    starting_time = datetime.fromtimestamp(int(from_ts), timezone(timedelta(hours=5, minutes=30))).strftime("%Y-%m-%d %H:%M:%S %z IST")
    
    return f"Duration {duration.strip()} starting from {starting_time}"


def fmt_prefix_2(raw_value):
    """Format volume using binary prefixes (1024-based), matching webtrisul fmt_prefix_2."""
    units_str = [
        (1099511627776.0, "T", "%.2f %s"),
        (1073741824.0, "G", "%.2f %s"),
        (1048576.0, "M", "%.2f %s"),
        (1024.0, "K", "%.2f %s"),
        (1, "", "%d %s"),
        (1e-03, "m", "%.3f %s"),
        (1e-06, "u", "%.3f %s"),
        (1e-09, "n", "%.3f %s"),
        (1e-12, "p", "%.3f %s"),
    ]
    if raw_value == 0 or raw_value is None:
        return "0"
    for threshold, unit, fmt in units_str:
        if raw_value >= threshold:
            return (fmt % (raw_value / threshold, unit)).strip()
    return str(raw_value)


def fmt_prefix_10(raw_value):
    """Format bandwidth using decimal prefixes (1000-based), matching webtrisul fmt_prefix_10."""
    units_str = [
        (1e12, "T", "%.2f %s"),
        (1e9, "G", "%.2f %s"),
        (1e6, "M", "%.2f %s"),
        (1000, "K", "%.2f %s"),
        (1, "", "%d %s"),
        (1e-03, "m", "%.3f %s"),
        (1e-06, "u", "%.3f %s"),
        (1e-09, "n", "%.3f %s"),
        (1e-12, "p", "%.3f %s"),
    ]
    if raw_value == 0 or raw_value is None:
        return "0"
    for threshold, unit, fmt in units_str:
        if raw_value >= threshold:
            return (fmt % (float(raw_value) / threshold, unit)).strip()
    return str(raw_value)


def fmt_volume(val, units=""):
    """Format cumulative volume totals."""
    return fmt_prefix_2(val) + units.replace("ps", "")


def fmt_bw(val, units="bps"):
    """Format rate/bandwidth values."""
    prefix = fmt_prefix_10(float(val))
    units = units.lower()
    if prefix == "0":
        return f"0{units}"
    return f"{prefix} {units}"


def _stats_array_values(stats_array, meter_ids):
    """Extract per-meter values from a StatsArray protobuf message."""
    if not stats_array or not stats_array.values:
        return [0] * len(meter_ids)
    vals = list(stats_array.values)
    return [int(vals[mid]) if mid < len(vals) and vals[mid] is not None else 0 for mid in meter_ids]


def _resolve_meter_ids(counter_group_guid, meters, zmq_endpoint):
    """Resolve meter names or numeric IDs to meter indices and metadata."""
    cg_info = countergroup_info(zmq_endpoint, get_meter_info=True)
    if "error" in cg_info:
        raise ValueError(cg_info["error"])

    group = None
    for g in cg_info.get("groupDetails", []):
        if g.get("guid", "").upper() == counter_group_guid.upper():
            group = g
            break
    if not group:
        raise ValueError(f"Counter group GUID {counter_group_guid} not found")

    meter_lookup = {}
    for m in group.get("meters", []):
        mid = int(m.get("id", 0))
        meter_lookup[mid] = {
            "id": mid,
            "name": m.get("name", ""),
            "description": m.get("description", ""),
            "units": m.get("units", "bps"),
        }

    resolved = []
    for meter in meters:
        meter_str = str(meter).strip()
        if meter_str.isdigit():
            mid = int(meter_str)
            if mid not in meter_lookup:
                raise ValueError(f"Meter ID {mid} not found in counter group {group.get('name', counter_group_guid)}")
            resolved.append(meter_lookup[mid])
            continue

        needle = meter_str.lower().replace(" ", "")
        match = None
        for mid, info in meter_lookup.items():
            candidates = [
                info.get("description", ""),
                info.get("name", ""),
                str(mid),
            ]
            for candidate in candidates:
                c = candidate.lower().replace(" ", "")
                if c == needle or needle in c or c in needle:
                    match = info
                    break
            if match:
                break
        if not match:
            available = [f"{m['id']}:{m['description'] or m['name']}" for m in meter_lookup.values()]
            raise ValueError(f"Meter '{meter}' not found. Available meters: {', '.join(available)}")
        resolved.append(match)

    return resolved, group


def _get_key_meter_stats(counter_group_guid, key, meter_ids, from_ts, to_ts, zmq_endpoint, meters_info):
    """Fetch total/min/max/avg/latest stats for a key, matching custom key monitor logic."""
    req = trp_pb2.Message()
    req.trp_command = req.COUNTER_ITEM_NG_REQUEST
    ng = req.counter_item_ng_request
    ng.counter_group = counter_group_guid
    ng.key.label = str(key).strip().lower()
    ng.volumes_only = 1
    getattr(ng.time_interval, "from").tv_sec = int(from_ts)
    ng.time_interval.to.tv_sec = int(to_ts)

    resp = get_response(zmq_endpoint, req)
    duration_secs = max(int(to_ts) - int(from_ts), 1)

    if resp.HasField("rate_volumes") and resp.rate_volumes.values:
        totals = _stats_array_values(resp.rate_volumes, meter_ids)
    elif resp.HasField("totals") and resp.totals.values:
        totals = _stats_array_values(resp.totals, meter_ids)
    else:
        totals = [0] * len(meter_ids)

    maximums = _stats_array_values(resp.maximums, meter_ids) if resp.HasField("maximums") else [0] * len(meter_ids)
    minimums = _stats_array_values(resp.minimums, meter_ids) if resp.HasField("minimums") else [0] * len(meter_ids)
    latests = _stats_array_values(resp.latests, meter_ids) if resp.HasField("latests") else [0] * len(meter_ids)
    samples = _stats_array_values(resp.samples, meter_ids) if resp.HasField("samples") else [1] * len(meter_ids)

    averages = []
    for i, mid in enumerate(meter_ids):
        unit = meters_info[mid].get("units", "bps")
        val = totals[i] or 0
        if unit.lower().endswith("ps"):
            averages.append(int(val / duration_secs) if duration_secs > 0 else val)
        else:
            samp = samples[i] or 1
            averages.append(int(val / samp) if samp else val)

    for i, mid in enumerate(meter_ids):
        if meters_info[mid].get("units", "") == "Bps":
            maximums[i] = (maximums[i] or 0) * 8
            minimums[i] = (minimums[i] or 0) * 8
            latests[i] = (latests[i] or 0) * 8
            averages[i] = (averages[i] or 0) * 8

    key_label = resp.key.label if resp.HasField("key") and resp.key.label else str(key)
    return {
        "key": key_label,
        "meters": {
            mid: {
                "name": meters_info[mid]["description"] or meters_info[mid]["name"],
                "units": meters_info[mid]["units"],
                "totals": totals[i],
                "maximums": maximums[i],
                "minimums": minimums[i],
                "averages": averages[i],
                "latests": latests[i],
            }
            for i, mid in enumerate(meter_ids)
        },
    }


_IST = timezone(timedelta(hours=5, minutes=30))


def _format_excel_value(value, fmt="text", fmt_args=None):
    """Format a cell value for Excel export."""
    fmt_args = fmt_args or {}
    if fmt in (None, "", "text"):
        return "" if value is None else value
    if fmt in ("util_pct", "percent"):
        return _fmt_util_pct(value)
    if fmt == "number":
        if value is None:
            return ""
        decimals = int(fmt_args.get("decimals", 2))
        try:
            return round(float(value), decimals)
        except (TypeError, ValueError):
            return value
    if fmt == "volume":
        return fmt_volume(value, fmt_args.get("units", ""))
    if fmt in ("bandwidth", "bw"):
        return fmt_bw(value, fmt_args.get("units", "bps"))
    if fmt == "datetime_epoch":
        if value is None:
            return ""
        try:
            return datetime.fromtimestamp(
                int(value),
                _IST,
            ).strftime(fmt_args.get("strftime", "%Y-%m-%d %H:%M:%S %z"))
        except (TypeError, ValueError, OSError):
            return str(value)
    return "" if value is None else value


def _normalize_excel_columns(columns):
    """Normalize column definitions to a consistent list of dicts."""
    normalized = []
    for col in columns:
        if isinstance(col, str):
            normalized.append({"header": col, "key": col, "format": "text", "format_args": {}, "width": None})
        elif isinstance(col, dict):
            header = col.get("header") or col.get("key") or col.get("name")
            key = col.get("key") or header
            if not header:
                raise ValueError("Each column must have a 'header' (or 'key').")
            normalized.append({
                "header": header,
                "key": key,
                "format": col.get("format", "text"),
                "format_args": col.get("format_args") or {},
                "width": col.get("width"),
            })
        else:
            raise ValueError(f"Invalid column definition: {col}")
    if not normalized:
        raise ValueError("At least one column is required.")
    return normalized


def _excel_row_values(row, columns, row_is_list=False):
    """Extract formatted cell values for one data row."""
    values = []
    for i, col in enumerate(columns):
        if row_is_list:
            raw = row[i] if i < len(row) else None
        else:
            raw = row.get(col["key"]) if isinstance(row, dict) else None
        values.append(_format_excel_value(raw, col["format"], col["format_args"]))
    return values


def _apply_excel_column_merges(ws, data_start_row, merge_col_indexes, row_count):
    """Vertically merge consecutive rows; empty cells continue the previous group."""
    from openpyxl.styles import Alignment

    for col_idx in merge_col_indexes:
        group_start = None
        for offset in range(row_count):
            row_num = data_start_row + offset
            cell_val = ws.cell(row=row_num, column=col_idx).value
            if cell_val not in (None, ""):
                if group_start is not None and row_num - group_start > 1:
                    ws.merge_cells(
                        start_row=group_start, start_column=col_idx,
                        end_row=row_num - 1, end_column=col_idx,
                    )
                    ws.cell(row=group_start, column=col_idx).alignment = Alignment(vertical="top")
                group_start = row_num
        if group_start is not None:
            end_row = data_start_row + row_count - 1
            if end_row > group_start:
                ws.merge_cells(
                    start_row=group_start, start_column=col_idx,
                    end_row=end_row, end_column=col_idx,
                )
                ws.cell(row=group_start, column=col_idx).alignment = Alignment(vertical="top")


def _auto_size_excel_columns(ws, min_width=10, max_width=40, skip_columns=None):
    """Auto-size worksheet columns based on cell content."""
    skip_columns = skip_columns or set()
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col_letter in skip_columns:
            continue
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _build_excel_report(
    columns,
    rows,
    title=None,
    metadata=None,
    from_ts=None,
    to_ts=None,
    filename=None,
    sheet_name="Report",
    merge_columns=None,
    include_generated_timestamp=True,
    header_bold=True,
    output_dir="/tmp",
):
    """Write a flexible Excel report with customizable columns, rows, and formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    columns = _normalize_excel_columns(columns)
    if not rows:
        raise ValueError("At least one data row is required.")

    row_is_list = isinstance(rows[0], (list, tuple))

    filepath = filename or f"excel_report_{int(datetime.now().timestamp())}.xlsx"
    if not filepath.startswith("/"):
        filepath = f"{output_dir.rstrip('/')}/{filepath}"
    if not filepath.endswith(".xlsx"):
        filepath += ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = str(sheet_name)[:31]

    if title:
        ws.append([title])
    if from_ts is not None and to_ts is not None:
        ws.append([epoch_to_duration(from_ts, to_ts)])
    for line in metadata or []:
        if line is not None and str(line).strip():
            ws.append([str(line)])
    if include_generated_timestamp:
        now_str = datetime.now(_IST).strftime("%Y-%m-%d %H:%M:%S %z")
        ws.append([f"Generated at {now_str}"])
    if title or (from_ts is not None and to_ts is not None) or metadata or include_generated_timestamp:
        ws.append([])

    header_row = ws.max_row + 1
    headers = [col["header"] for col in columns]
    ws.append(headers)
    if header_bold:
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=header_row, column=col_idx).font = Font(bold=True)

    data_start_row = header_row + 1
    for row in rows:
        ws.append(_excel_row_values(row, columns, row_is_list=row_is_list))

    key_to_index = {col["key"]: i + 1 for i, col in enumerate(columns)}
    merge_col_indexes = [
        key_to_index[mk] for mk in (merge_columns or []) if mk in key_to_index
    ]
    if merge_col_indexes:
        _apply_excel_column_merges(ws, data_start_row, merge_col_indexes, len(rows))

    fixed_width_columns = set()
    for i, col in enumerate(columns):
        if col.get("width"):
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = col["width"]
            fixed_width_columns.add(col_letter)
    _auto_size_excel_columns(ws, skip_columns=fixed_width_columns)

    wb.save(filepath)
    return filepath


def _build_key_monitor_excel(
    title,
    from_ts,
    to_ts,
    rows,
    filename,
):
    """Write the custom key monitor Excel report."""
    flat_rows = []
    for key_name, meter_rows in rows:
        for i, meter_row in enumerate(meter_rows):
            flat_rows.append({
                "name": key_name if i == 0 else "",
                "meter": meter_row["meter"],
                "total": meter_row["total"],
                "max": meter_row["max"],
                "min": meter_row["min"],
                "avg": meter_row["avg"],
                "latest": meter_row["latest"],
            })

    return _build_excel_report(
        columns=[
            {"header": "Name", "key": "name"},
            {"header": "Meter", "key": "meter"},
            {"header": "Total", "key": "total"},
            {"header": "Max", "key": "max"},
            {"header": "Min", "key": "min"},
            {"header": "Avg", "key": "avg"},
            {"header": "Latest", "key": "latest"},
        ],
        rows=flat_rows,
        title=title,
        from_ts=from_ts,
        to_ts=to_ts,
        filename=filename,
        sheet_name="Key Monitor",
        merge_columns=["name"],
    )


def _fmt_util_pct(val):
    """Format a utilization gauge value as a percentage string."""
    if val is None or val < 0:
        return "-"
    return f"{round(float(val), 2)}%"


def _key_attrs_to_dict(keyt):
    """Convert KeyT attributes to a name->value dict."""
    if not keyt or not keyt.attributes:
        return {}
    return {a.attr_name: a.attr_value for a in keyt.attributes}


def _is_system_key(key):
    """Return True for Trisul aggregate/system keys that are not real entities."""
    if not key:
        return True
    key = str(key)
    return key in ("Others", "SYS:GROUP_TOTALS") or key.startswith("SYS:")


def _is_interface_key(key):
    """FlowIntfs keys are routerKey_ifindex (e.g. AC.1B.08.01_0000000B)."""
    return bool(key) and not _is_system_key(key) and "_" in str(key)


def _get_time_window(zmq_endpoint, duration_secs, start_ts=None, end_ts=None):
    """Return (from_ts, to_ts) for a report window."""
    req = trp_pb2.Message()
    req.trp_command = req.TIMESLICES_REQUEST
    req.time_slices_request.get_total_window = True
    tint_resp = get_response(zmq_endpoint, req)
    from_ts_val = int(start_ts) if start_ts else int(tint_resp.total_window.to.tv_sec) - int(duration_secs)
    to_ts_val = int(end_ts) if end_ts else int(tint_resp.total_window.to.tv_sec)
    return from_ts_val, to_ts_val


def _fetch_flowintf_topper(zmq_endpoint, meter, max_count, from_ts, to_ts):
    """Fetch FlowIntfs topper keys with attributes for the given window."""
    req = trp_pb2.Message()
    req.trp_command = req.COUNTER_GROUP_TOPPER_REQUEST
    topper = req.counter_group_topper_request
    topper.counter_group = FLOWINTFS_GUID
    topper.meter = int(meter)
    topper.maxitems = int(max_count)
    topper.get_key_attributes = True
    topper.inverse_key_filter = "SYS:GROUP"
    getattr(topper.time_interval, "from").tv_sec = int(from_ts)
    topper.time_interval.to.tv_sec = int(to_ts)
    resp = get_response(zmq_endpoint, req)
    keys = [k for k in resp.keys if _is_interface_key(k.key)]
    logging.info(
        f"[_fetch_flowintf_topper] meter={meter} raw={len(resp.keys)} "
        f"interfaces={len(keys)}"
    )
    return keys


def _fetch_router_names(router_keys, zmq_endpoint):
    """Resolve router key -> display name from FlowGens."""
    if not router_keys:
        return {}
    req = trp_pb2.Message()
    req.trp_command = req.SEARCH_KEYS_REQUEST
    q = req.search_keys_request
    q.counter_group = FLOWGENS_GUID
    q.keys.extend(list(router_keys))
    q.maxitems = len(router_keys)
    resp = get_response(zmq_endpoint, req)
    lookup = {}
    for k in resp.keys:
        name = k.label.strip() if k.label and k.label.strip() else k.readable
        lookup[k.key] = name
    return lookup


def _build_pnb_excel(title, from_ts, to_ts, rows, filename):
    """Write the PNB interface utilization Excel report."""
    return _build_excel_report(
        columns=[
            {"header": "Router IP", "key": "router_ip"},
            {"header": "Router Name", "key": "router_name"},
            {"header": "Interface", "key": "interface"},
            {"header": "Interface Description", "key": "interface_description"},
            {"header": "In Utilization", "key": "in_utilization"},
            {"header": "Out Utilization", "key": "out_utilization"},
            {"header": "Total Utilization", "key": "total_utilization"},
        ],
        rows=rows,
        title=title,
        from_ts=from_ts,
        to_ts=to_ts,
        filename=filename,
        sheet_name="PNB Interface Utilization",
    )


# TRP Helper functions

_RESPONSE_FIELD_MAP = {
    'COUNTER_GROUP_INFO_RESPONSE': 'counter_group_info_response',
    'TIMESLICES_RESPONSE': 'time_slices_response',
    'COUNTER_GROUP_TOPPER_RESPONSE': 'counter_group_topper_response',
    'COUNTER_ITEM_RESPONSE': 'counter_item_response',
    'QUERY_ALERTS_RESPONSE': 'query_alerts_response',
    'QUERY_SESSIONS_RESPONSE': 'query_sessions_response',
    'SEARCH_KEYS_RESPONSE': 'search_keys_response',
}

def unwrap_response(data):
    try:
        logging.info("[unwrap_response] Unwrapping response")
        resp = trp_pb2.Message()
        resp.ParseFromString(data)
        
        command_name = None
        for x in resp.DESCRIPTOR.enum_types:
            val = x.values_by_number.get(int(resp.trp_command))
            if val:
                command_name = val.name
                break
        
        logging.info(f"[unwrap_response] Response command: {command_name}")
        
        field_name = _RESPONSE_FIELD_MAP.get(command_name)
        if field_name and resp.HasField(field_name):
            return getattr(resp, field_name)
        
        return resp
    except Exception as e:
        logging.error(f"[unwrap_response] Error unwrapping response: {str(e)}")
        raise


def get_response(zmq_endpoint, req, timeout_ms=10000):
    socket = None
    try:
        logging.info(f"[get_response] Connecting to {zmq_endpoint}")
        zmq_context = _global_zmq_context
        socket = zmq_context.socket(zmq.REQ)
        socket.setsockopt(zmq.LINGER, 0)
        socket.setsockopt(zmq.RCVTIMEO, timeout_ms)
        socket.setsockopt(zmq.SNDTIMEO, timeout_ms)
        
        socket.connect(zmq_endpoint)
        socket.send(req.SerializeToString())
        
        data = socket.recv()
        logging.info(f"[get_response] Received {len(data)} bytes")
        return unwrap_response(data)
    except zmq.Again:
        error_msg = f"[get_response] ZMQ timeout after {timeout_ms}ms - no response from {zmq_endpoint}"
        logging.error(error_msg)
        raise Exception(error_msg)
    except zmq.ZMQError as e:
        error_msg = f"[get_response] ZMQ error: {str(e)}"
        logging.error(error_msg)
        raise Exception(error_msg)
    except Exception as e:
        error_msg = f"[get_response] Error: {str(e)}"
        logging.error(error_msg)
        raise
    finally:
        if socket:
            try:
                socket.close()
            except Exception as e:
                logging.warning(f"Error closing socket: {str(e)}")





# TRP tools

@mcp.tool()
def list_all_available_counter_groups(context: str = "context0", zmq_endpoint: str = None):
    """List all available counter groups from Trisul via ZMQ for a given context or the zmq_endpoint.
    Arguments: 
        context (str): Context name, should be like context_XYZ or default or context0 etc.
        zmq_endpoint (str): ZMQ endpoint in the format "tcp://<ip_address>:<port>", for example "tcp://10.16.8.44:5008". The IP address and port may vary.
    Returns: dict: Dictionary containing counter group information.
    Example: list_all_available_counter_groups("context_XYZ") or list_all_available_counter_groups("tcp://10.16.8.44:5008") -> 
        {
            "groupDetails": [
                {"guid": "{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}", "name": "ABC"},
                {"guid": "{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}", "name": "XYZ"},
                ...
            ]
        }
    """
    try:
        if not zmq_endpoint:
            context = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{context}/run/trp_0"

        logging.info(f"[list_all_available_counter_groups] Listing all available counter groups for zmq_endpoint: {zmq_endpoint}")

        all_cgs = countergroup_info(zmq_endpoint, get_meter_info=False)
        
        if "error" in all_cgs:
            logging.error(f"[list_all_available_counter_groups] Error from countergroup_info: {all_cgs['error']}")
            return json_to_toon({"error": all_cgs["error"], "groupDetails": []})
        
        group_details = all_cgs.get("groupDetails", [])
        logging.info(f"[list_all_available_counter_groups] Processing {len(group_details)} counter groups")
        
        simplified_groups = []
        for g in group_details:
            try:
                simplified_groups.append({"guid": g["guid"], "name": g["name"]})
            except KeyError as e:
                logging.warning(f"[list_all_available_counter_groups] Missing key in group details: {str(e)}, skipping group")
                continue
        
        logging.info(f"[list_all_available_counter_groups] Retrieved {len(simplified_groups)} counter groups")
        return json_to_toon({"groupDetails": simplified_groups})
        
    except Exception as e:
        logging.error(f"[list_all_available_counter_groups] Error in list_all_available_counter_groups: {str(e)}", exc_info=True)
        return json_to_toon({"error": str(e), "groupDetails": []})



@mcp.tool()
def get_cginfo_from_countergroup_name(countergroup_name: str, context: str = "context0", zmq_endpoint: str = None):
    """Fetch counter group details by counter group name from Trisul via ZMQ for a given context or the zmq_endpoint.
    and it will also fetch meters info for each counter group so that we can determine what each meter means and its index.
    for example if we want to get the counter group guid for "ABCDE" and meter index for "Received traffic" we can use this function.
    Arguments: 
        countergroup_name (str): Counter group name
        context (str): Context name, should be like context_XYZ or default or context0 etc.
        zmq_endpoint (str): ZMQ endpoint in the format "tcp://<ip_address>:<port>", for example "tcp://10.16.8.44:5008". The IP address and port may vary.
    Returns: dict: Counter Group Details . If not found, it will return the list of all available counter groups name and the guid.
    Example: get_cginfo_from_countergroup_name("ABC", "context0") -> 
        {
            "guid": "{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}","name": "ABCDE","bucketSize": "60000","topperBucketSize": "300",
            "timeInterval": { "from": {"tvSec": "1718711400","tvUsec": "0"}, "to": {"tvSec": "1718712060","tvUsec": "0"} },
            "meters": [
                { "id": 0, "type": "VT_RATE_COUNTER", "topcount": 1000, "name": "Bps", "description": "Total", "units": "Bps" },
                { "id": 1, "type": "VT_RATE_COUNTER", "topcount": 1000, "name": "Bps", "description": "Recv", "units": "Bps" },
                { "id": 2, "type": "VT_RATE_COUNTER", "topcount": 1000, "name": "Bps", "description": "Xmit", "units": "Bps" }
            ]
        }
    """
    try:
        countergroup_name = str(countergroup_name)
        if not zmq_endpoint:
            context = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{context}/run/trp_0"
            
        logging.info(f"[get_cginfo_from_countergroup_name] Fetching counter group info for name: {countergroup_name}, zmq_endpoint: {zmq_endpoint}")
        
        # Get all counter groups (with meter info if available)
        all_cgs = countergroup_info(zmq_endpoint, get_meter_info=True)
        
        if "error" in all_cgs:
            logging.error(f"[get_cginfo_from_countergroup_name] Error from countergroup_info: {all_cgs['error']}")
            return json_to_toon({"name": countergroup_name, "guid": f"Error: {all_cgs['error']}"})
        
        group_details = all_cgs.get("groupDetails", [])
        logging.info(f"[get_cginfo_from_countergroup_name] Retrieved {len(group_details)} counter groups")
        
        group_names = []
        normalized_search_name = countergroup_name.lower().replace(" ", "")
        logging.info(f"[get_cginfo_from_countergroup_name] Normalized search name: {normalized_search_name}")
        
        for group in group_details:
            try:
                group_name = group.get("name", "")
                group_names.append(group_name)
                
                normalized_group_name = group_name.lower().replace(" ", "")
                if normalized_group_name == normalized_search_name:
                    logging.info(f"[get_cginfo_from_countergroup_name] Found matching counter group: {group_name}")
                    return json_to_toon(group)  # return full raw group dict
            except Exception as e:
                logging.warning(f"[get_cginfo_from_countergroup_name] Error processing group: {str(e)}, skipping")
                continue
        
        # If not found
        logging.warning(f"[get_cginfo_from_countergroup_name] Counter group '{countergroup_name}' not found. Available groups: {group_names}")
        return json_to_toon({
            "name": countergroup_name,
            "guid": "Not Found",
            "available_groups": group_names
        })
            
    except Exception as e:
        logging.error(f"[get_cginfo_from_countergroup_name] Error in get_cginfo_from_countergroup_name: {str(e)}", exc_info=True)
        return json_to_toon({"name": countergroup_name, "guid": f"Error: {str(e)}"})
    


@mcp.tool()
def get_counter_group_topper(counter_group_guid: str, meter: int = 0, duration_secs: int = 3600, max_count: int = 10, context: str = "context0", zmq_endpoint: str = None):
    """
    Fetch the topper metrics for a given counter group and meter over the last `duration_secs` seconds.
    Arguments: 
    counter_group_guid (str): GUID of the Counter group , meter (int): Meter index, duration_secs (int): Duration in seconds, max_count (int): maximum number of toppers retrive, 
    context (str): Context name, 
    zmp_endpoint (str): ZMQ endpoint in the format "tcp://<ip_address>:<port>", for example "tcp://10.16.8.44:5008". The IP address and port may vary.
    Returns: dict: Dictionary containing topper metrics.
    Example: get_counter_group_topper("{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}", 0, 3600, "context0") or
             get_counter_group_topper("{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}", 0, 3600, "tcp://10.16.8.44:5008")-> 
    {'counterGroup': '{889900CC-0063-11A5-8380-FEBDBABBDBEA}', 'meter': '0', 'keys': 
    [key': '0A.19.1E.97', 'readable': '10.25.30.151', 'label': '10.25.30.151', 'description': '', 'metric': '242287', 'metricMax': '137112', 'metricMin': '105175', 'metricAvg': '121143'}, 
    {'key': '0A.1A.0C.68', 'readable': '10.26.12.104', 'label': '10.26.12.104', 'description': '', 'metric': '227337', 'metricMax': '227337', 'metricMin': '227337', 'metricAvg': '227337'}]}
    """
    
    global _global_zmq_context
    socket = None
    
    try:
        counter_group_guid = str(counter_group_guid)
        meter = int(meter)
        duration_secs = int(duration_secs)
        max_count = int(max_count)
        if not zmq_endpoint:
            context = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{context}/run/trp_0"
        
        logging.info(f"[get_counter_group_topper] Fetching counter group topper: counter_group_guid={counter_group_guid}, meter={meter}, duration_secs={duration_secs}, max_count={max_count}, context={zmq_endpoint}")

        # Step 1: Get available timeslices
        logging.info("[get_counter_group_topper] Step 1: Getting available timeslices")
        req = trp_pb2.Message()
        req.trp_command = req.TIMESLICES_REQUEST
        req.time_slices_request.get_total_window = True
        resp = get_response(zmq_endpoint, req)
        logging.info("[get_counter_group_topper] Timeslices received")

        # Step 2: Build topper request
        logging.info("[get_counter_group_topper] Step 2: Building topper request")
        req = trp_pb2.Message()
        req.trp_command = req.COUNTER_GROUP_TOPPER_REQUEST
        req.counter_group_topper_request.counter_group = counter_group_guid
        req.counter_group_topper_request.meter = meter
        req.counter_group_topper_request.maxitems = max_count

        # Step 3: Time interval for last duration_secs
        logging.info("[get_counter_group_topper] Step 3: Setting time interval")
        tm = trp_pb2.TimeInterval()
        tm.to.tv_sec = resp.total_window.to.tv_sec
        object = getattr(tm, 'from')
        object.tv_sec = tm.to.tv_sec - duration_secs
        req.counter_group_topper_request.time_interval.MergeFrom(tm)
        logging.info(f"[get_counter_group_topper] Time interval: from={object.tv_sec}, to={tm.to.tv_sec}")

        # Step 4: Get topper response
        logging.info("[get_counter_group_topper] Step 4: Getting topper response")
        resp = get_response(zmq_endpoint, req)
        logging.info("[get_counter_group_topper] Successfully retrieved counter group topper")

        # Step 5: Return JSON-serializable dict
        return json_to_toon(MessageToDict(resp))
    
    except Exception as e:
        logging.error(f"[get_counter_group_topper] Error in get_counter_group_topper: {str(e)}", exc_info=True)
        return json_to_toon({"error": str(e)})



@mcp.tool()
def get_key_traffic_data(counter_group: str, readable: Any = None, duration_secs: int = 3600, start_ts: int = None, end_ts: int = None, context: str = "context0", zmq_endpoint: str = None):
    """
    Fetch the key traffic metrics for a given counter group and readable over the last `duration_secs` seconds.
    the duration_secs can be any value other than 0.
    It will return data for all meter.
    But it will not Generate the chart display the data. you need to call the next appropriate tool to do that.
    Arguments: 
        counter_group (str): Counter group GUID, readable (str): Key value, duration_secs (int): Duration in seconds, 
        context (str): Context name, 
        zmq_endpoint (str): ZMQ endpoint in the format "tcp://<ip_address>:<port>", for example "tcp://10.16.8.44:5008". The IP address and port may vary.
    Returns: dict: Dictionary containing key traffic metrics.
    always try to pass the readable value as readable format like 10.25.46.1 or https, not in key format like 0A.19.2E.01 or p-01BB.
    Example: key_traffic("{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}", "163.70.151.21", 3600, "XYZ") or
        key_traffic("{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}", "163.70.151.21", 1748409542, 1748412428, "XYZ") or
        key_traffic("{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}", "163.70.151.21", 1748409542, 1748412428, "tcp://10.16.8.44:5008") ->
    output:
        {
            "counterGroup": "{XXXXXXXX-XXXX-XXXXXXXX-XXXXXXXXXXXXX}",
            "key": { "key": "A3.46.97.15", "readable": "163.70.151.21", "label": "163.70.151.21", "description": ""},
            "stats": [
                {
                    "tsTvSec": "1718711760",
                    "values": [ "302793", "5328", "297465", "281", "25", "0", "0", "302793", "0", "0", "21", "0", "0", "0", "0", "0", "67", "0", "0"]
                },
                {
                    "tsTvSec": "1718711820",
                    "values": ["253915","5819","248097","246","18","0","0","253915","0","0","20","0","0","0","0","0","53","0","0" ]
                }
            ]
        }
    """
    
    global _global_zmq_context
    socket = None
    
    try:        
        counter_group = str(counter_group)
        if readable:
            readable = str(readable)
        if duration_secs is not None:
            duration_secs = int(duration_secs)
        if start_ts is not None:
            start_ts = int(start_ts)
        if end_ts is not None:
            end_ts = int(end_ts)
        
        if not zmq_endpoint:
            context = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{context}/run/trp_0"
 
        logging.info(f"[get_key_traffic_data] Fetching key traffic: counter_group={counter_group}, readable={readable}, duration_secs={duration_secs}, start_ts={start_ts}, end_ts={end_ts}, zmq_endpoint={zmq_endpoint}")
                        

        #Construct time request
        try:
            logging.info("[get_key_traffic_data] Constructing TIMESLICES_REQUEST")
            req = trp_pb2.Message()
            req.trp_command = req.TIMESLICES_REQUEST
            req.time_slices_request.get_total_window = True
            logging.info("[get_key_traffic_data] Sending TIMESLICES_REQUEST")
            tint_resp = get_response(zmq_endpoint, req)
            logging.info("[get_key_traffic_data] Received timeslices response")
        except Exception as e:
            logging.error(f"[get_key_traffic_data] Error getting timeslices: {str(e)}")
            raise


        #construct counter item request request for internal host
        try:
            logging.info("[get_key_traffic_data] Constructing COUNTER_ITEM_REQUEST")
            req = trp_pb2.Message()
            req.trp_command = req.COUNTER_ITEM_REQUEST
            req.counter_item_request.counter_group = counter_group
            req.counter_item_request.key.label = readable.lower()
            logging.info(f"[get_key_traffic_data] Counter item request configured: counter_group={counter_group}, readable={readable}")
        except Exception as e:
            logging.error(f"[get_key_traffic_data] Error constructing counter item request: {str(e)}")
            raise

        #construct time interval for last 1 hour
        try:
            logging.info("[get_key_traffic_data] Constructing time interval")
            tm = trp_pb2.TimeInterval()
            tm.MergeFrom(tint_resp.total_window)
            object = getattr(tm, 'from')
            object.tv_sec = tm.to.tv_sec - duration_secs
            
            logging.info(f"[get_key_traffic_data] Default time interval: from={object.tv_sec}, to={tm.to.tv_sec}")

            #assign time interval to counter group topper request
            if start_ts and end_ts:
                logging.info(f"[get_key_traffic_data] Overriding time interval with start_ts={start_ts}, end_ts={end_ts}")
                object = getattr(tm, 'from')
                object.tv_sec = start_ts
                object = getattr(tm, 'to')
                object.tv_sec = end_ts
                logging.info(f"[get_key_traffic_data] Time interval set: from={start_ts}, to={end_ts}")
            else:
                logging.info(f"[get_key_traffic_data] Time interval set: from={object.tv_sec}, to={tm.to.tv_sec} (duration: {duration_secs}s)")
                
            req.counter_item_request.time_interval.MergeFrom(tm)
        except Exception as e:
            logging.error(f"[get_key_traffic_data] Error setting time interval: {str(e)}")
            raise
        
        logging.info("[get_key_traffic_data] Sending COUNTER_ITEM_REQUEST")
        resp = get_response(zmq_endpoint, req)
        logging.info("[get_key_traffic_data] Successfully received key traffic response")
        
        result = MessageToDict(resp)
        logging.info(f"[get_key_traffic_data] Response converted to dict, keys: {result.keys()}")
        
        return json_to_toon(result)
    
        
    except zmq.ZMQError as e:
        logging.error(f"[get_key_traffic_data] ZMQ error in key_traffic: {str(e)}", exc_info=True)
        return json_to_toon({"error": f"ZMQ error: {str(e)}"})
    except Exception as e:
        logging.error(f"[get_key_traffic_data] Error in key_traffic: {str(e)}", exc_info=True)
        return json_to_toon({"error": str(e)})
    finally:
        if socket:
            try:
                socket.close()
                logging.info("[get_key_traffic_data] Final socket cleanup")
            except Exception as e:
                logging.warning(f"[get_key_traffic_data] Error in final socket cleanup: {str(e)}")



@mcp.tool()
def get_alerts_data(
    alert_group: str,
    duration_secs: int = 3600,
    start_ts: int = None,
    end_ts: int = None,
    context: str = "context0",
    zmq_endpoint: str = None,
    maxitems: int = 100,
    group_by_fieldname: str = None,
    resolve_keys: bool = True,
    approx_count_only: bool = False,
    source_ip: Any = None,
    destination_ip: Any = None,
    source_port: Any = None,
    destination_port: Any = None,
    any_ip: Any = None,
    any_port: Any = None,
    ip_pair: List[str] = None,
    sigid: Any = None,
    classification: Any = None,
    priority: Any = None,
    aux_message1: Any = None,
    aux_message2: Any = None,
    message_regex: Any = None,
    idlist: List[str] = None
):
    """
    Retrieve alert telemetry from Trisul using QUERY_ALERTS_REQUEST.

    This MCP tool acts as an alert intelligence fetcher that surfaces curated alert
    records for a specified Alert Group. It supports flexible temporal scoping, granular
    field-level filtering, grouping, and regex-based matching to enable downstream
    enrichment, correlation, or analytics workloads.

    **Mandatory Requirement:**
        - `alert_group` must be a valid Trisul Alert Group GUID.
          The function will not operate if a non-GUID or malformed value is supplied.

    Parameters:
        alert_group (str): REQUIRED. Trisul Alert Group GUID. Must be a valid GUID string.
        duration_secs (int): Relative time window in seconds. Ignored if start_ts and end_ts are provided.
        start_ts (int): Start of absolute time window (epoch seconds). Must be paired with end_ts.
        end_ts (int): End of absolute time window (epoch seconds). Must be paired with start_ts.
        context (str): Trisul context identifier. Defaults to "context0".
        zmq_endpoint (str): Custom TRP ZMQ endpoint. Auto-computed if omitted.
        maxitems (int): Hard ceiling on number of alert records returned.
        group_by_fieldname (str): Field to group output by (e.g. "sigid", "source_ip").
        resolve_keys (bool): Resolve internal keys into readable fields.
        approx_count_only (bool): Return approximate counts only, without full alert details.

        # Filters (always use readable values, not internal key format)
        source_ip (str): Filter by source IP.
        destination_ip (str): Filter by destination IP.
        source_port (str): Filter by source port.
        destination_port (str): Filter by destination port.
        any_ip (str): Match either source or destination IP.
        any_port (str): Match either source or destination port.
        ip_pair (list): One or multiple [src_ip, dst_ip] filter pairs.
        sigid (str): Filter by signature ID.
        classification (str): Filter by alert classification.
        priority (str): Filter by alert priority.
        aux_message1 (str): Text match against dispatch message field 1.
        aux_message2 (str): Text match against dispatch message field 2.
        message_regex (str): Regex match on alert message payload.
        idlist (list): Retrieve specific alerts by ID.

    Returns:
        dict: Parsed Trisul alert response, including raw or grouped alert intelligence payload.

    Usage Notes:
        - Always provide a valid GUID for `alert_group` to avoid request rejection.
        - Absolute time window (start_ts/end_ts) overrides duration_secs if both provided.
        - Optimized for downstream dashboards, analytics engines, and correlation pipelines.
    """


    try:
        logging.info(f"[get_alerts_data] Start | alert_group={alert_group} duration_secs={duration_secs} start_ts={start_ts} end_ts={end_ts}")

        if not zmq_endpoint:
            ctx = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{ctx}/run/trp_0"
        logging.info(f"[get_alerts_data] ZMQ endpoint: {zmq_endpoint}")


        logging.info("[get_alerts_data] Requesting TIMESLICES window")
        req = trp_pb2.Message()
        req.trp_command = req.TIMESLICES_REQUEST
        req.time_slices_request.get_total_window = True
        req.time_slices_request.get_total_window = True
        tint_resp = get_response(zmq_endpoint, req)

        tm = trp_pb2.TimeInterval()
        tm.MergeFrom(tint_resp.total_window)
        getattr(tm, 'from').tv_sec = tm.to.tv_sec - duration_secs

        if start_ts and end_ts:
            getattr(tm, 'from').tv_sec = int(start_ts)
            getattr(tm, 'to').tv_sec = int(end_ts)
            logging.info(f"[get_alerts_data] Custom time window applied: {start_ts} to {end_ts}")

        req = trp_pb2.Message()
        req.trp_command = req.QUERY_ALERTS_REQUEST
        q = req.query_alerts_request

        q.alert_group = alert_group
        q.time_interval.MergeFrom(tm)
        q.maxitems = int(maxitems)
        q.resolve_keys = bool(resolve_keys)
        q.approx_count_only = bool(approx_count_only)

        if group_by_fieldname:
            q.group_by_fieldname = group_by_fieldname
            logging.info(f"[get_alerts_data] Group by: {group_by_fieldname}")

        def set_keyt(field, val):
            if val is None:
                return
            getattr(q, field).label = str(val).lower()
            logging.info(f"[get_alerts_data] Filter applied: {field}={val}")

        set_keyt('source_ip', source_ip)
        set_keyt('destination_ip', destination_ip)
        set_keyt('source_port', source_port)
        set_keyt('destination_port', destination_port)
        set_keyt('any_ip', any_ip)
        set_keyt('any_port', any_port)
        set_keyt('sigid', sigid)
        set_keyt('classification', classification)
        set_keyt('priority', priority)

        if aux_message1:
            logging.info(f"[get_alerts_data] aux_message1={aux_message1}")
            q.aux_message1 = aux_message1
        if aux_message2:
            logging.info(f"[get_alerts_data] aux_message2={aux_message2}")
            q.aux_message2 = aux_message2
        if message_regex:
            logging.info(f"[get_alerts_data] message_regex={message_regex}")
            q.message_regex = message_regex

        if idlist:
            q.idlist.extend([str(x) for x in idlist])
            logging.info(f"[get_alerts_data] idlist count={len(idlist)}")

        if ip_pair:
            pairs = ip_pair
            if isinstance(pairs, list) and pairs and isinstance(pairs[0], str):
                pairs = [pairs]
            for p in pairs:
                if len(p) != 2:
                    logging.warning(f"[get_alerts_data] Invalid ip_pair skipped: {p}")
                    continue
                kt1 = q.ip_pair.add()
                kt1.label = str(p[0]).lower()
                kt2 = q.ip_pair.add()
                kt2.label = str(p[1]).lower()
            logging.info(f"[get_alerts_data] ip_pair count={len(pairs)}")

        logging.info("[get_alerts_data] Executing QUERY_ALERTS_REQUEST")
        resp = get_response(zmq_endpoint, req)
        
        return json_to_toon(MessageToDict(resp))

    except Exception as e:
        logging.error(f"[get_alerts_data] Error: {str(e)}", exc_info=True)
        return json_to_toon({"error": str(e)})



@mcp.tool()
def get_flows_or_sessions_data(
        session_group: str = "{99A78737-4B41-4387-8F31-8077DB917336}",
        key: Any = None,
        source_ip: Any = None,
        source_port: Any = None,
        dest_ip: Any = None,
        dest_port: Any = None,
        any_ip: Any = None,
        any_port: Any = None,
        ip_pair: List[str] = None,
        protocol: Any = None,
        flowtag: Any = None,
        nf_routerid: Any = None,
        nf_ifindex_in: Any = None,
        nf_ifindex_out: Any = None,
        subnet_24: Any = None,
        subnet_16: Any = None,
        maxitems: int = 100,
        volume_filter: int = 0,
        resolve_keys: bool = True,
        outputpath: str = None,
        idlist: List[str] = None,
        any_nf_ifindex: Any = None,
        duration_secs: int = 60,
        start_ts: int = None,
        end_ts: int = None,
        context: str = "context0",
        zmq_endpoint: str = None
    ):
    """
    Unified QuerySessions API pull.

    Business Value:
        Single entry point to query Trisul sessions with multi-criteria filtering.
        All fields in QuerySessionsRequest are supported through `filters`.
        Fields provided are implicitly AND-ed, enabling precision flow slicing.

    Args:
        session_group: Session group GUID. Default is main Flow Tracker.
        key: Match a Trisul internal session key.
        source_ip, dest_ip: Match flow endpoints by IP.
        source_port, dest_port: Match L4 ports.
        any_ip, any_port: Match either source or destination.
        ip_pair: List of 2 IPs. Matches flows between the pair.
        protocol: L4 protocol (6=TCP,17=UDP,1=ICMP).
        flowtag: Match flow tag text.
        nf_routerid: NetFlow router ID.
        nf_ifindex_in, nf_ifindex_out: NetFlow interface filters.
        subnet_24, subnet_16: Match flows inside subnet ranges.
        maxitems: Max records returned. Default 200.
        volume_filter: Only return flows > X bytes.
        resolve_keys: Resolve keys to readable format.
        outputpath: Write results to hub as CSV instead of returning.
        idlist: Flow IDs to retrieve directly. Skips filters.
        any_nf_ifindex: Match IN or OUT NF interface.
        duration_secs: Time window if timestamps not provided.
        start_ts, end_ts: Epoch timestamps override duration_secs.
        context: Trisul context.
        zmq_endpoint: Custom TRP endpoint.
        
        filters (dict): Maps directly to QuerySessionsRequest fields.
                        Examples:
                        {
                            "any_ip": "10.1.1.1",
                            "source_ip": "192.168.1.5",
                            "dest_port": "443",
                            "protocol": "6",
                            "flowtag": "malware",
                            "ip_pair": ["10.1.1.1","8.8.8.8"],
                            "subnet_24": "172.16.5.0"
                        }

    Returns:
        dict: Parsed sessions response as Python dict.
    """

    try:
        if not zmq_endpoint:
            context = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{context}/run/trp_0"

        logging.info(f"[QuerySessions] TRP endpoint={zmq_endpoint}")
            
        # Step 1: Pull Time Window
        req = trp_pb2.Message()
        req.trp_command = req.TIMESLICES_REQUEST
        req.time_slices_request.get_total_window = True
        req.time_slices_request.get_total_window = True
        tint_resp = get_response(zmq_endpoint, req)

        tm = trp_pb2.TimeInterval()
        tm.MergeFrom(tint_resp.total_window)
        
        if not start_ts or not end_ts:
            duration_secs = int(duration_secs)
            start_ts = tm.to.tv_sec - duration_secs
            end_ts = tm.to.tv_sec
        
        start_ts = int(start_ts)
        end_ts = int(end_ts)
        
        getattr(tm, 'from').tv_sec = start_ts
        getattr(tm, 'to').tv_sec = end_ts

        # Step 2: Build QuerySessionsRequest
        req = trp_pb2.Message()
        req.trp_command = req.QUERY_SESSIONS_REQUEST
        q = req.query_sessions_request

        q.session_group = session_group
        q.time_interval.MergeFrom(tm)
        q.maxitems = maxitems
        q.volume_filter = volume_filter
        q.resolve_keys = resolve_keys
        if outputpath: q.outputpath = outputpath

        if key: q.key = str(key)
        if source_ip: q.source_ip.label = str(source_ip)
        if source_port: q.source_port.label = str(source_port)
        if dest_ip: q.dest_ip.label = str(dest_ip)
        if dest_port: q.dest_port.label = str(dest_port)
        if any_ip: q.any_ip.label = str(any_ip)
        if any_port: q.any_port.label = str(any_port)
        if protocol: q.protocol.label = str(protocol)
        if flowtag: q.flowtag = str(flowtag)
        if nf_routerid: q.nf_routerid.label = str(nf_routerid)
        if nf_ifindex_in: q.nf_ifindex_in.label = str(nf_ifindex_in)
        if nf_ifindex_out: q.nf_ifindex_out.label = str(nf_ifindex_out)
        if subnet_24: q.subnet_24 = str(subnet_24)
        if subnet_16: q.subnet_16 = str(subnet_16)
        if any_nf_ifindex: q.any_nf_ifindex.label = str(any_nf_ifindex)

        if ip_pair and len(ip_pair) == 2:
            p1 = q.ip_pair.add(); p1.label = str(ip_pair[0])
            p2 = q.ip_pair.add(); p2.label = str(ip_pair[1])

        if idlist:
            for fid in idlist:
                q.idlist.append(fid)

        logging.info(f"[QuerySessions] Executing QuerySessions with provided filters")
        
        resp = MessageToDict(get_response(zmq_endpoint, req))
        
        resp["sessions"][:] = [
            s for s in resp.get("sessions", [])
            if int(s["timeInterval"]["from"]["tvSec"]) <= end_ts and int(s["timeInterval"]["to"]["tvSec"]) >= start_ts
        ][-maxitems:]
        
        logging.info(resp)
        
        return json_to_toon(resp)
        
    except Exception as e:
        logging.error(f"[QuerySessions] Exception: {e}")
        return json_to_toon({"error": str(e)})





# Non TRP tools

@mcp.tool()
def create_crosskey_counter_group( context: str = "context0", name: str = None, description: str = "No description", toppers_interval: int = 300, bucket_size: int = 60, track_hi_water: int = 500, track_lo_water: int = 100, tail_prune_factor: int = None, last_topper_bucket_ts: str = None, row_status: str = "Active", cardinality_estimate_bits: int = None, topper_traffic_only: bool = None, enable_slice_keys: int = 1, resolver_counter_guid: str = None, cross_guid1: str = None, cross_guid2: str = None, cross_guid3: str = None, balance_depth : int = None):
    """
    Create a new crosskey counter group in Trisul.
    We cannot create the crosskey with the zmq_endpoint, it require the context name.
    Arguments:
        context (str): Context name, should be like context_XYZ or default or context0 etc.
        name (str): Name of the counter group
        description (str): Description of the counter group (Default: "No description")
        topn_commit_interval_secs (int): Time interval for toppers traffic in seconds (Default: 300)
        bucket_size (int): Time Interval for key traffic in seconds (Default: 60)
        track_hi_water (int): High water mark for tracking (Default: 500)
        track_lo_water (int): Low water mark for tracking (Default: 100)
        tail_prune_factor (int): Tail prune factor (Default: None)
        row_status (str): Counter group status(enabled or disabled), e.g., "Active" (Default: "Active")
        cardinality_estimate_bits (int): Cardinality estimate bits (Default: None)
        topper_traffic_only (bool): Whether to track topper traffic only or key traffic also (Default: None)
        enable_slice_keys (bool): Whether to enable slice keys (Default: True)
        resolver_counter_guid (str): Resolver counter GUID (Default: None)
        Returns: dict: Dictionary with details of the created counter group or error message.
        """
    
    conn = None
    cursor = None
    
    try:
        logging.info(f"[create_crosskey_counter_group] Creating crosskey counter group: name={name}, context={context}")
        
        if not name:
            error_msg = "[create_crosskey_counter_group] Counter group name is required"
            logging.error(error_msg)
            return {"status": "error", "message": error_msg}
        
        context = normalize_context(context)
        db_path = f"/usr/local/var/lib/trisul-config/domain0/{context}/profile0/TRISULCONFIG.SQDB"
        logging.info(f"[create_crosskey_counter_group] Connecting to database: {db_path}")
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        logging.info("[create_crosskey_counter_group] Database connection established")
        
        # Create the Counter Group
        cg_sql = """
            INSERT INTO TRISUL_COUNTER_GROUPS
            (CounterGUID, Name, Description, TopNCommitIntervalSecs, BucketSizeMS, TrackHiWater, TrackLoWater, TailPruneFactor, LastTopperBucketTS, RowStatus, CardinalityEstimateBits, TopperTrafficOnly, EnableSliceKeys, CreateTimestamp, ModifyTimestamp, ResolverCounterGUID)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        
        new_guid = f'{{{str(uuid.uuid4()).upper()}}}'
        logging.info(f"[create_crosskey_counter_group] Generated new GUID: {new_guid}")
        
        cg_values = (
            new_guid,
            name,
            description,
            toppers_interval,
            bucket_size * 1000,
            track_hi_water,
            track_lo_water,
            tail_prune_factor,
            last_topper_bucket_ts,
            row_status,
            cardinality_estimate_bits,
            topper_traffic_only,
            enable_slice_keys,
            int(datetime.datetime.now().timestamp()),
            int(datetime.datetime.now().timestamp()),
            resolver_counter_guid
        )

        logging.info(f"[create_crosskey_counter_group] Executing counter group insert with values: {cg_values}")
        cursor.execute(cg_sql, cg_values)
        conn.commit()
        logging.info("[create_crosskey_counter_group] Counter group inserted successfully")
        
        cross_sql = """
            INSERT INTO TRISUL_COUNTER_GROUP_CROSSKEYS
            (CounterGUID, ParentCounterGUID, CrosskeyCounterGUID, CrosskeyThirdCounterGUID,
            KeyLength1, KeyLength2, KeyLength3, BalanceDepth1, BalanceDepth2)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        cross_values = (
            new_guid,
            cross_guid1,
            cross_guid2,
            cross_guid3,
            None, 
            None, 
            None, 
            balance_depth, 
            None
        )
        
        logging.info(f"[create_crosskey_counter_group] Executing crosskey insert with values: {cross_values}")
        cursor.execute(cross_sql, cross_values)
        conn.commit()
        logging.info("[create_crosskey_counter_group] Crosskey configuration inserted successfully")
        
        success_msg = f"[create_crosskey_counter_group] Counter group '{name}' successfully created with guid {new_guid}."
        logging.info(success_msg)
        return {"status": "success", "message": success_msg}
        
    except sqlite3.IntegrityError as e:
        error_msg = f"[create_crosskey_counter_group] Database integrity error: {str(e)}"
        logging.error(error_msg)
        if conn:
            conn.rollback()
        return {"status": "error", "message": error_msg}
    except sqlite3.OperationalError as e:
        error_msg = f"[create_crosskey_counter_group] Database operational error: {str(e)}"
        logging.error(error_msg)
        return {"status": "error", "message": error_msg}
    except Exception as e:
        error_msg = f"[create_crosskey_counter_group] Error creating counter group: {str(e)}"
        logging.error(error_msg, exc_info=True)
        if conn:
            conn.rollback()
        return {"status": "error", "message": error_msg}
    finally:
        if cursor:
            try:
                cursor.close()
                logging.info("[create_crosskey_counter_group] Cursor closed")
            except Exception as e:
                logging.warning(f"[create_crosskey_counter_group] Error closing cursor: {str(e)}")
        if conn:
            try:
                conn.close()
                logging.info("[create_crosskey_counter_group] Database connection closed")
            except Exception as e:
                logging.warning(f"[create_crosskey_counter_group] Error closing database connection: {str(e)}")



@mcp.tool()
def rag_query(question: str):
    """
    Perform a RAG (Retrieval-Augmented Generation) query using Gemini and ChromaDB.
    It does not need any context or the zmq_endpoint
    Arguments: question (str): The question to query.
    Returns: str: The answer generated by Gemini.
        Example: rag_query("what is crosskey?") -> 
        "Crosskey is a feature in Trisul that allows you to combine multiple counter groups to create a new composite counter group. 
        For example, you can create a crosskey counter group that combines the 'Source IP' and 'Destination IP' counter groups to track traffic between specific IP pairs."
    """

    try:
        logging.info(f"[rag_query] Starting RAG query for question: {question}")
        
        # Embed query
        # Embed query
        logging.info("[rag_query] Initializing Embedding Model via LLMFactory")
        try:
            env_path = Path(__file__).resolve().parent / ".env"
            factory = LLMFactory(env_path=env_path, logging=logging)
            embedding_model = factory.get_embedding_llm()
            
            if not embedding_model:
                 return "Error: Embedding model not configured or API key missing. Please configure it using the CLI."

            logging.info(f"[rag_query] Generating embedding for question using {factory.embedding_model}")
            q_emb = embedding_model.embed_query(question)
            logging.info(f"[rag_query] Embedding generated successfully, dimension: {len(q_emb)}")
            
        except Exception as e:
            logging.error(f"[rag_query] Error generating embedding: {str(e)}", exc_info=True)
            return f"Error: Failed to generate embedding - {str(e)}"

        # Search in Chroma
        try:
            logging.info("[rag_query] Initializing ChromaDB client")
            CHROMA_STORE = Path(__file__).resolve().parent / "chroma_store"
            logging.info(f"[rag_query] ChromaDB store path: {CHROMA_STORE}")
            
            if not CHROMA_STORE.exists():
                logging.warning(f"[rag_query] ChromaDB store path does not exist: {CHROMA_STORE}")
            
            chroma_client = chromadb.PersistentClient(path=str(CHROMA_STORE))
            collection = chroma_client.get_or_create_collection("pdf_docs")
            logging.info("[rag_query] ChromaDB client initialized successfully")
        except Exception as e:
            logging.error(f"[rag_query] Error initializing ChromaDB: {str(e)}", exc_info=True)
            return f"Error: Failed to initialize ChromaDB - {str(e)}"

        try:
            logging.info("[rag_query] Querying ChromaDB collection")
            results = collection.query(
                query_embeddings=[q_emb], 
                n_results=3, 
                include=['documents', 'distances', 'embeddings']
            )
            logging.info("[rag_query] ChromaDB query completed successfully")
            logging.info(f"[rag_query] Query results structure - Keys: {results.keys()}")
            
            if 'documents' in results:
                logging.info(f"[rag_query] Number of document groups: {len(results['documents'])}")
                if results['documents']:
                    logging.info(f"[rag_query] Number of documents in first group: {len(results['documents'][0])}")
            
            if 'distances' in results:
                logging.info(f"[rag_query] Distances: {results.get('distances', [])}")
        except Exception as e:
            logging.error(f"[rag_query] Error querying ChromaDB: {str(e)}", exc_info=True)
            return f"Error: Failed to query ChromaDB - {str(e)}"

        try:
            logging.info("[rag_query] Extracting retrieved documents")
            if 'documents' not in results or not results['documents']:
                logging.warning("[rag_query] No documents found in query results")
                return "No relevant documents found in the knowledge base."
            
            retrieved_docs = results["documents"][0]
            logging.info(f"[rag_query] Retrieved {len(retrieved_docs)} documents")
            
            if not retrieved_docs:
                logging.warning("[rag_query] Retrieved documents list is empty")
                return "No relevant documents found in the knowledge base."
            
            for i, doc in enumerate(retrieved_docs):
                logging.info(f"[rag_query] Document {i+1} preview: {doc[:100]}..." if len(doc) > 100 else f"Document {i+1}: {doc}")
        except (KeyError, IndexError) as e:
            logging.error(f"[rag_query] Error extracting documents from results: {str(e)}")
            return f"Error: Failed to extract documents - {str(e)}"
        except Exception as e:
            logging.error(f"[rag_query] Unexpected error processing documents: {str(e)}", exc_info=True)
            return f"Error: Failed to process documents - {str(e)}"
        
        # Build prompt
        try:
            logging.info("[rag_query] Building context from retrieved documents")
            context = "\n".join(retrieved_docs)
            logging.info(f"[rag_query] Context built successfully, length: {len(context)} characters")
            logging.info(f"[rag_query] Context preview: {context[:200]}..." if len(context) > 200 else f"Context: {context}")
            
            return context
        except Exception as e:
            logging.error(f"[rag_query] Error building context: {str(e)}", exc_info=True)
            return f"Error: Failed to build context - {str(e)}"
            
    except Exception as e:
        logging.error(f"[rag_query] Unexpected error in rag_query: {str(e)}", exc_info=True)
        return f"Error: An unexpected error occurred - {str(e)}"






# UI related tools

@mcp.tool()
def show_table(data: dict):
    """
    Displays a structured table in the Chat UI. 
    Use this tool when the user explicitly asks for a table or when presenting topper rankings in a structured format for Web UI/API consumers.
    
    **MANDATORY**: Whenever you use this tool, you must also provide a textual summary stating the source Counter Group name, and the Key, Label, and Readable attributes of the entity being displayed. If label and readable are same, show only one.
    
    Args:
        data (dict): Table configuration and data.
                     Example format:
                     {
                        "title": "Top 5 Applications",
                        "headers": ["Application", "Total Traffic", "Flows"],
                        "rows": [
                            ["HTTPS", "1.52 GB", "120"],
                            ["HTTP", "850 MB", "45"],
                            ["DNS", "12 MB", "300"]
                        ]
                     }
    """
    logging.info(f"[show_table] Preparing table for Chat UI integration")
    
    # Validate the input data
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except Exception:
            logging.error("[show_table] Invalid table data format. Expected dict or JSON string.")
            return {"status": "error", "message" : "Invalid table data format from LLM"}
            
    return {"status": "success", "message": "Table data prepared for UI display."}


@mcp.tool()
def show_line_chart(data, save_image: bool = False):
    """
    Plots a static traffic chart (line chart) using matplotlib based on the provided JSON-like input and show it in a new pop-up window.
    the input values should be in raw bytes format  not in mb or kb.
    the time stamps should be in epoc seconds format as integer not in the string date time format like this '2025-10-16 01:00:00'.
    It does not need any context name or the zmq_endpoint.
    It can be used to display the traffic data for any key like IP address, protocol, port etc over a time period.
    
    **MANDATORY**: Whenever you use this tool, you must also provide a textual summary stating the source Counter Group name, and the Key, Label, and Readable attributes of the entity being displayed. If label and readable are same, show only one.
    
    Args:
        data (dict): Line Chart configuration and series data.
                     Example format:
                     {
                        "title": "Network Traffic Over 24 Hours", "x_label": "Time", "y_label": "Traffic",
                        "keys": [
                            {
                                "timestamps": [1718714100, 1718714160], "legend_label": "Upload", "color": "red", "values": [32432, 37293]
                            },
                            ...
                        ]
                     }
        save_image (bool): set it 'True' to save the chart as an image file and don't display it in pop-up window. set it False to display the chart in pop-up window. Default is False.
    """
    
    logging.info(f"[show_line_chart] Generating the line chart for the given data")
    
        
    # Validate the input data
    if isinstance(data, str):
        try:
            data = ast.literal_eval(data)
        except Exception:
            try:
                data = json.loads(data)
            except Exception:
                logging.error("[show_line_chart] Invalid line chart data format. Expected dict or JSON string.")
                return {"status": "error", "message" : "Invalid line chart data format from LLM", "message_to_llm" : "Call this mcp tool (show_line_chart) again with the valid line chart data format. don't retry this more than 3 times in a row", "file_path": None}
    else:
        data = dict(data)
        
    # Check if the length of timestamps and values are same
    for series in data.get("keys", []):
        if(len(series["timestamps"]) != len(series["values"])):
            logging.error("[show_line_chart] Invalid line chart data format. The length of timestamps and values should be same.")
            return {"status": "error", "message" : "Invalid line chart data format from LLM, the length of timestamps and values should be same", "message_to_llm" : "Call this mcp tool (show_line_chart) again with the valid line chart data format. don't retry this more than 3 times in a row", "file_path": None}


    # Save the chart as an image file if save_image is True
    if(save_image):
        file_path = f"/tmp/line_chart_{int(datetime.now().timestamp())}_{random.randint(1000, 9999)}.png"
        logging.info(f"[show_line_chart] save_image is set to True, so saving the chart as an image file instead of displaying it. path: {file_path}")
        return {"status": "success", "message" : f"The line chart is saved as an image file successfully.", "file_path": file_path}
    else:
        return {"status": "success", "message" : "The line chart has been generated and is being displayed in the UI.", "file_path": None}



@mcp.tool()
def show_pie_chart(data, save_image: bool = False):
    """
    Plots a static traffic chart (pie chart) using matplotlib based on the provided JSON-like input and show it in a new pop-up window.
    
    Usage:
        To display pie chart to show the topper values for any counter group and meter.
        It does not need any context name or the zmq_endpoint.
        It can be used to display the traffic distribution for any key like IP address, protocol, port etc.

    **MANDATORY**: Whenever you use this tool, you must also provide a textual summary stating the source Counter Group name, and the Key, Label, and Readable attributes of the entity being displayed. If label and readable are same, show only one.

    Args:
        data (dict): Pie Chart configuration and series data.
                    Example format:
                    {
                       'chart_title': 'Top Applications by Traffic',
                       'legend_title': 'Applications',
                       'labels': ['HTTP', 'HTTPS', 'DNS', 'SSH', 'FTP'],
                       'volumes': [5776124, 4733635, 1028367, 14143, 14001],
                       'colors': ['#1F77B4', '#FF7F0E', '#2CA02C', '#D62728', '#9467BD']
                    }
        save_image (bool): set it 'True' to save the chart as an image file and don't display it in pop-up window. set it False to display the chart in pop-up window. Default is False.

    Returns:
        dict: Status and message about the pie chart display.
    """
    logging.info(f"[show_pie_chart] Generating the pie chart for the given data")
    
    
    # Validate the input data
    if isinstance(data, str):
        try:
            data = ast.literal_eval(data)
        except Exception:
            try:
                data = json.loads(data)
            except Exception:
                logging.error("[show_pie_chart] Invalid pie chart data format. Expected dict or JSON string.")
                return {"status": "error", "message" : "Invalid pie chart data format from LLM", "message_to_llm" : "Call this mcp tool (show_pie_chart) again with the valid pie chart data format. don't retry this more than 3 times in a row", "file_path": None}
    else:
        data = dict(data)


    data["volumes"] = [
        eval(str(v)) if isinstance(v, str) and "*" in v else v
        for v in data.get("volumes", [])
    ]
    
    if(save_image):
        file_path = f"/tmp/pie_chart_{int(datetime.now().timestamp())}_{random.randint(1000, 9999)}.png"
        logging.info(f"[show_pie_chart] save_image is set to True, so saving the chart as an image file instead of displaying it. path: {file_path}")
        return {"status": "success", "message" : f"The pie chart is saved as an image file successfully.", "file_path": file_path}
    else:
        logging.info(f"[show_pie_chart] save_image is set to False, so displaying the chart in the UI.")
        return {"status": "success", "message" : "The pie chart has been generated and is being displayed in the UI.", "file_path": None}



@mcp.tool()
def generate_key_monitor_excel_report(
    counter_group_guid: str,
    keys: List[str],
    meters: List[str],
    title: str = "Key Monitor Report",
    duration_secs: int = 86400,
    start_ts: int = None,
    end_ts: int = None,
    filename: str = None,
    context: str = "context0",
    zmq_endpoint: str = None,
):
    """
    Generate an Excel report for specific keys and meters in a counter group,
    matching the Custom Key Monitor format in webtrisul UI.

    Provide three inputs:
      1. counter_group_guid - GUID of the counter group (e.g. FlowIntfs, Hosts, Apps)
      2. keys - list of key labels/readable names to include (e.g. ["10.25.30.151", "cloud-Standard-PC-"])
      3. meters - list of meter names or numeric IDs (e.g. ["Total", "Received", "Transmit"] or ["0", "1", "2"])

    The report columns are: Name, Meter, Total, Max, Min, Avg, Latest — with the key name
    shown once per group (rowspan) and one row per meter, exactly like the web UI.

    Args:
        counter_group_guid (str): Counter group GUID.
        keys (list[str]): Keys to fetch (use readable labels, not internal hex keys).
        meters (list[str]): Meter descriptions/names or numeric meter IDs.
        title (str): Report title shown at the top of the Excel sheet.
        duration_secs (int): Lookback window in seconds (default 86400 = 1 day). Ignored if start_ts/end_ts given.
        start_ts (int): Optional absolute start epoch seconds.
        end_ts (int): Optional absolute end epoch seconds.
        filename (str): Output filename (saved under /tmp/). Auto-generated if omitted.
        context (str): Trisul context (default "context0").
        zmq_endpoint (str): Custom TRP ZMQ endpoint. Auto-computed if omitted.

    Returns:
        dict: status, message, and file_path of the generated .xlsx file.

    Example:
        generate_key_monitor_excel_report(
            counter_group_guid="{889900CC-0063-11A5-8380-FEBDBABBDBEA}",
            keys=["cloud-Standard-PC-", "102.104.119.80.rev."],
            meters=["Total", "Received", "Transmit"],
            title="demo title",
            duration_secs=86400,
        )
    """
    try:
        counter_group_guid = str(counter_group_guid).strip()
        if not keys:
            return {"status": "error", "message": "At least one key is required.", "file_path": None}
        if not meters:
            return {"status": "error", "message": "At least one meter is required.", "file_path": None}

        if not zmq_endpoint:
            ctx = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{ctx}/run/trp_0"

        logging.info(
            f"[generate_key_monitor_excel_report] guid={counter_group_guid} "
            f"keys={keys} meters={meters} duration_secs={duration_secs}"
        )

        req = trp_pb2.Message()
        req.trp_command = req.TIMESLICES_REQUEST
        req.time_slices_request.get_total_window = True
        tint_resp = get_response(zmq_endpoint, req)

        from_ts_val = int(start_ts) if start_ts else int(tint_resp.total_window.to.tv_sec) - int(duration_secs)
        to_ts_val = int(end_ts) if end_ts else int(tint_resp.total_window.to.tv_sec)

        resolved_meters, _ = _resolve_meter_ids(counter_group_guid, meters, zmq_endpoint)
        meter_ids = [m["id"] for m in resolved_meters]
        meters_info = {m["id"]: m for m in resolved_meters}

        report_rows = []
        for key in keys:
            try:
                stats = _get_key_meter_stats(
                    counter_group_guid, key, meter_ids, from_ts_val, to_ts_val, zmq_endpoint, meters_info
                )
            except Exception as ex:
                logging.warning(f"[generate_key_monitor_excel_report] Failed for key '{key}': {ex}")
                stats = {
                    "key": str(key),
                    "meters": {
                        mid: {
                            "name": meters_info[mid]["description"] or meters_info[mid]["name"],
                            "units": meters_info[mid]["units"],
                            "totals": 0, "maximums": 0, "minimums": 0, "averages": 0, "latests": 0,
                        }
                        for mid in meter_ids
                    },
                }

            meter_rows = []
            sort_total = 0
            for mid in meter_ids:
                m = stats["meters"][mid]
                units = m["units"]
                if mid == meter_ids[0]:
                    sort_total = m["totals"]
                meter_rows.append({
                    "meter": m["name"],
                    "total": fmt_volume(m["totals"]),
                    "max": fmt_bw(m["maximums"], units),
                    "min": fmt_bw(m["minimums"], units),
                    "avg": fmt_bw(m["averages"], units),
                    "latest": fmt_bw(m["latests"], units),
                })
            report_rows.append((stats["key"], meter_rows, sort_total))

        report_rows.sort(key=lambda item: item[2], reverse=True)
        report_rows = [(key, meter_rows) for key, meter_rows, _ in report_rows]

        if not filename:
            safe_title = "".join(c if c.isalnum() or c in "-_" else "_" for c in title)[:40]
            filename = f"key_monitor_{safe_title}_{int(datetime.now().timestamp())}.xlsx"

        filepath = _build_key_monitor_excel(title, from_ts_val, to_ts_val, report_rows, filename)

        logging.info(f"[generate_key_monitor_excel_report] Report saved to {filepath}")
        return {
            "status": "success",
            "message": f"Excel report generated successfully at {filepath}",
            "file_path": filepath,
            "keys_count": len(report_rows),
            "meters": [m["description"] or m["name"] for m in resolved_meters],
            "duration": epoch_to_duration(from_ts_val, to_ts_val),
        }

    except Exception as e:
        logging.error(f"[generate_key_monitor_excel_report] Error: {e}", exc_info=True)
        return {"status": "error", "message": str(e), "file_path": None}


@mcp.tool()
def generate_pnb_excel_report(
    max_count: int = 25,
    duration_secs: int = 3600,
    sort_meter: int = 4,
    title: str = "PNB Interface Utilization Report",
    filename: str = None,
    start_ts: int = None,
    end_ts: int = None,
    context: str = "context0",
    zmq_endpoint: str = None,
):
    """
    Generate the PNB Excel report for top N router interfaces by utilization.

    The report contains one row per interface with columns:
    Router IP, Router Name, Interface, Interface Description,
    In Utilization, Out Utilization, Total Utilization.

    Interfaces are ranked by Total Utilization (average of In and Out utilization)
    over the selected time window. Data is sourced from the FlowIntfs counter group.

    Args:
        max_count (int): Number of top interfaces to include (default 25).
        duration_secs (int): Lookback window in seconds (default 3600 = 1 hour).
            Ignored if start_ts/end_ts are provided.
        sort_meter (int): Initial FlowIntfs meter used to fetch candidate interfaces
            (default 4 = Recv/In Utilization).
        title (str): Report title shown at the top of the Excel sheet.
        filename (str): Output filename (saved under /tmp/). Auto-generated if omitted.
        start_ts (int): Optional absolute start epoch seconds.
        end_ts (int): Optional absolute end epoch seconds.
        context (str): Trisul context (default "context0").
        zmq_endpoint (str): Custom TRP ZMQ endpoint. Auto-computed if omitted.

    Returns:
        dict: status, message, file_path, and summary metadata.

    Example:
        generate_pnb_excel_report(max_count=25, duration_secs=3600)
    """
    try:
        max_count = int(max_count)
        duration_secs = int(duration_secs)
        sort_meter = int(sort_meter)
        if max_count <= 0:
            return {"status": "error", "message": "max_count must be greater than 0.", "file_path": None}

        if not zmq_endpoint:
            ctx = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{ctx}/run/trp_0"

        logging.info(
            f"[generate_pnb_excel_report] max_count={max_count} duration_secs={duration_secs} "
            f"sort_meter={sort_meter}"
        )

        from_ts_val, to_ts_val = _get_time_window(zmq_endpoint, duration_secs, start_ts, end_ts)
        candidate_count = min(max(max_count * 3, max_count), 500)
        topper_keys = _fetch_flowintf_topper(
            zmq_endpoint, sort_meter, candidate_count, from_ts_val, to_ts_val
        )

        # Utilization meters may be empty; fall back to total-traffic topper for candidates.
        if not topper_keys and sort_meter != 0:
            logging.info(
                "[generate_pnb_excel_report] No interfaces from util meter "
                f"{sort_meter}; falling back to meter 0"
            )
            topper_keys = _fetch_flowintf_topper(
                zmq_endpoint, 0, candidate_count, from_ts_val, to_ts_val
            )

        if not topper_keys:
            return {
                "status": "error",
                "message": (
                    "No interface data found for the selected time window. "
                    "SYS:GROUP aggregate keys were excluded."
                ),
                "file_path": None,
            }

        resolved_meters, _ = _resolve_meter_ids(
            FLOWINTFS_GUID, [4, 5], zmq_endpoint
        )
        meter_ids = [m["id"] for m in resolved_meters]
        meters_info = {m["id"]: m for m in resolved_meters}
        in_meter_id = meter_ids[0]
        out_meter_id = meter_ids[1] if len(meter_ids) > 1 else meter_ids[0]

        router_keys = {
            keyt.key.split("_")[0]
            for keyt in topper_keys
            if _is_interface_key(keyt.key)
        }
        router_names = _fetch_router_names(router_keys, zmq_endpoint)

        report_rows = []
        for keyt in topper_keys:
            if not _is_interface_key(keyt.key):
                continue

            router_key = keyt.key.split("_")[0]
            readable = keyt.readable or keyt.key
            router_ip = readable.split("_")[0] if "_" in readable else router_key
            router_name = router_names.get(router_key, router_ip)

            attrs = _key_attrs_to_dict(keyt)
            interface = keyt.label.strip() if keyt.label and keyt.label.strip() else readable.split("_")[-1]
            interface_description = (
                attrs.get("snmp.ifalias")
                or (keyt.description.strip() if keyt.description else "")
                or "-"
            )

            lookup_key = keyt.readable or keyt.label or keyt.key
            try:
                stats = _get_key_meter_stats(
                    FLOWINTFS_GUID,
                    lookup_key,
                    meter_ids,
                    from_ts_val,
                    to_ts_val,
                    zmq_endpoint,
                    meters_info,
                )
                in_util = stats["meters"][in_meter_id]["averages"]
                out_util = stats["meters"][out_meter_id]["averages"]
            except Exception as ex:
                logging.warning(
                    f"[generate_pnb_excel_report] Failed for interface '{lookup_key}': {ex}"
                )
                in_util = -1
                out_util = -1

            total_util = (in_util + out_util) / 2 if in_util >= 0 and out_util >= 0 else -1
            report_rows.append({
                "router_ip": router_ip,
                "router_name": router_name,
                "interface": interface,
                "interface_description": interface_description,
                "in_utilization": _fmt_util_pct(in_util),
                "out_utilization": _fmt_util_pct(out_util),
                "total_utilization": _fmt_util_pct(total_util),
                "_sort_total_util": total_util,
            })

        report_rows.sort(key=lambda row: row["_sort_total_util"], reverse=True)
        report_rows = report_rows[:max_count]
        for row in report_rows:
            row.pop("_sort_total_util", None)

        if not report_rows:
            return {
                "status": "error",
                "message": "No interface utilization data could be collected.",
                "file_path": None,
            }

        if not filename:
            safe_title = "".join(c if c.isalnum() or c in "-_" else "_" for c in title)[:40]
            filename = f"pnb_interface_util_{safe_title}_{int(datetime.now().timestamp())}.xlsx"

        filepath = _build_pnb_excel(title, from_ts_val, to_ts_val, report_rows, filename)

        logging.info(f"[generate_pnb_excel_report] Report saved to {filepath}")
        return {
            "status": "success",
            "message": f"PNB Excel report generated successfully at {filepath}",
            "file_path": filepath,
            "interfaces_count": len(report_rows),
            "duration": epoch_to_duration(from_ts_val, to_ts_val),
            "columns": [
                "Router IP",
                "Router Name",
                "Interface",
                "Interface Description",
                "In Utilization",
                "Out Utilization",
                "Total Utilization",
            ],
        }

    except Exception as e:
        logging.error(f"[generate_pnb_excel_report] Error: {e}", exc_info=True)
        return {"status": "error", "message": str(e), "file_path": None}


@mcp.tool()
def generate_excel_report(
    columns: List[dict],
    rows: List[dict],
    title: str = None,
    metadata: List[str] = None,
    from_ts: int = None,
    to_ts: int = None,
    filename: str = None,
    sheet_name: str = "Report",
    merge_columns: List[str] = None,
    include_generated_timestamp: bool = True,
):
    """
    Generate a flexible Excel (.xlsx) report with custom columns, rows, and formatting.

    Use this as the generic report builder for any tabular Excel export. Specialized
    reports (key monitor, PNB, etc.) can also be built with this tool when you already
    have the data assembled.

    Args:
        columns (list[dict]): Column definitions. Each dict has:
            * header (str): Column header text (required)
            * key (str): Row dict key for this column (defaults to header)
            * format (str): Value formatter — one of:
                "text" (default), "percent"/"util_pct", "number", "volume",
                "bandwidth"/"bw", "datetime_epoch"
            * format_args (dict): Formatter options, e.g. {"decimals": 2}, {"units": "bps"}
            * width (int): Fixed column width in characters
        rows (list[dict]): Data rows; each dict is keyed by column "key" values.
        title (str): Optional report title shown at the top.
        metadata (list[str]): Optional extra header lines below the title.
        from_ts (int): Optional start epoch; adds a duration line when paired with to_ts.
        to_ts (int): Optional end epoch; adds a duration line when paired with from_ts.
        filename (str): Output filename (saved under /tmp/). Auto-generated if omitted.
        sheet_name (str): Excel worksheet name (default "Report", max 31 chars).
        merge_columns (list[str]): Column keys to vertically merge when consecutive
            rows leave the cell empty (useful for grouped/key-monitor style reports).
        include_generated_timestamp (bool): Add a "Generated at ..." line (default True).

    Returns:
        dict: status, message, file_path, row_count, and column headers.

    Examples:
        # Simple custom report
        generate_excel_report(
            title="Top Hosts",
            columns=[
                {"header": "Host", "key": "host"},
                {"header": "Total Bytes", "key": "bytes", "format": "volume"},
                {"header": "Avg Bandwidth", "key": "avg_bps", "format": "bandwidth"},
            ],
            rows=[
                {"host": "10.1.1.1", "bytes": 1048576, "avg_bps": 1500000},
                {"host": "10.1.1.2", "bytes": 524288, "avg_bps": 750000},
            ],
            from_ts=1718711400,
            to_ts=1718715000,
        )

        # Pre-formatted text values (no format conversion)
        generate_excel_report(
            columns=[
                {"header": "Name", "key": "name"},
                {"header": "Status", "key": "status"},
                {"header": "Value", "key": "value"},
            ],
            rows=[
                {"name": "router-a", "status": "OK", "value": "99.5%"},
                {"name": "router-b", "status": "WARN", "value": "82.1%"},
            ],
            sheet_name="Health Check",
        )
    """
    try:
        if not columns:
            return {"status": "error", "message": "At least one column is required.", "file_path": None}
        if not rows:
            return {"status": "error", "message": "At least one data row is required.", "file_path": None}

        normalized_columns = _normalize_excel_columns(columns)
        logging.info(
            f"[generate_excel_report] columns={len(normalized_columns)} rows={len(rows)} "
            f"sheet_name={sheet_name}"
        )

        if not filename:
            safe_title = "".join(
                c if c.isalnum() or c in "-_" else "_"
                for c in (title or sheet_name or "report")
            )[:40]
            filename = f"excel_{safe_title}_{int(datetime.now().timestamp())}.xlsx"

        filepath = _build_excel_report(
            columns=columns,
            rows=rows,
            title=title,
            metadata=metadata,
            from_ts=from_ts,
            to_ts=to_ts,
            filename=filename,
            sheet_name=sheet_name,
            merge_columns=merge_columns,
            include_generated_timestamp=include_generated_timestamp,
        )

        logging.info(f"[generate_excel_report] Report saved to {filepath}")
        if not os.path.isfile(filepath):
            return {
                "status": "error",
                "message": f"Excel file was not written to disk at {filepath}",
                "file_path": None,
            }
        return {
            "status": "success",
            "message": f"Excel report generated successfully at {filepath}",
            "file_path": filepath,
            "row_count": len(rows),
            "columns": [col["header"] for col in normalized_columns],
            "duration": epoch_to_duration(from_ts, to_ts) if from_ts is not None and to_ts is not None else None,
        }

    except Exception as e:
        logging.error(f"[generate_excel_report] Error: {e}", exc_info=True)
        return {"status": "error", "message": str(e), "file_path": None}


@mcp.tool()
def generate_trisul_report(pages, filename: str, report_title: str, from_ts, to_ts):
    """
    Generate a multi-page PDF report with multiple tables or  traffic charts (one per page).
    
    Args:
        filename (str): Output PDF file name.
        pages (list[dict]): Each dict = {'title': str, 'subtitle': str, 'data': list[list[str]]}
            Each page should have a title, subtitle, and data.  Data can be either a table or a chart.
            For table pages, 'data' is a 2D list representing rows and columns.
            For chart pages, 'data' is a dict with 'file_path' key pointing to the chart image file.
        report_title (str): Title of the report to be displayed in the header of all pages. The title should be short and descriptive within 2-4 words.
        from_ts (int): Start timestamp of the report duration (epoch seconds).
        to_ts (int): End timestamp of the report duration (epoch seconds).

    Example:
        pages = [
            {
                "type": "table",
                "title": "Internal Hosts",
                "subtitle": "Top internal hosts by total volume",
                "data": [
                    ["Internal Hosts", "Readable", "Flows", "Sent Bytes", "Received Bytes", "Total Bytes", "Percent"],
                    ["10.40.16.100", "10.40.16.100", "40", "49.61 K", "334.92 K", "384.53 K", "63.5"],
                    ["10.40.16.223", "10.40.16.223", "16", "28.98 K", "177.92 K", "206.90 K", "34.2"],
                ]
            },
            {
                "type": "chart",
                "title": "HTTPS Traffic Chart",
                "subtitle": "Showing HTTPS traffic trend over time",
                "file_path": "/tmp/traffic_chart_12345.png"
            },
    
        ]
        filename = 'trisul_https_report.pdf'
        report_title = 'HTTPS Traffic Report'
        from_ts = 1676610900
        to_ts = 1676614500
    """
    
    
    
    # Validate the input data
    if isinstance(pages, str):
        try:
            pages = ast.literal_eval(pages)
        except Exception:
            try:
                pages = json.loads(pages)
            except Exception:
                logging.error("[generate_trisul_report] Invalid pages data format from LLM.")
                return {"status": "error", "message" : "Invalid pages data format from LLM", "message_to_llm" : "Call this mcp tool (generate_trisul_report) again with the valid pages data format", "file_path": None}
    else:
        pages = list(pages)

    
    styles = getSampleStyleSheet()
    title_style = styles["Heading2"]
    title_style.leftIndent = 0
    title_style.spaceAfter = 10

    subtitle_style = styles["Heading5"]
    subtitle_style.leftIndent = 0
    subtitle_style.spaceAfter = 20
    subtitle_style.textColor = colors.HexColor("#800080")


    filename = f"/tmp/{filename}"
    
    pdf = SimpleDocTemplate(
        filename,
        pagesize=A4,
        leftMargin=10,
        rightMargin=15,
        topMargin=55,
        bottomMargin=70,
    )

    # Header/footer rendering
    def draw_header_footer(canvas, doc):
        width, height = A4

        # Header separator
        canvas.setStrokeColor(colors.black)
        canvas.line(15, height - 65, width - 15, height - 65)
        
        logo_path = Path(__file__).resolve().parent / "assets/logo_tlhs.png"
        
        duration_string = epoch_to_duration(from_ts, to_ts)
        
        
        # Logo
        try:
            canvas.drawImage(logo_path, 14, height - 63, width=69, height=49, mask='auto')
        except:
            pass

        # Header text
        canvas.setFillColorRGB(0, 0, 0)
        canvas.setFont("Helvetica", 14)
        canvas.drawRightString(width - 15, height - 28, report_title)
        canvas.setFont("Helvetica", 10)
        canvas.drawRightString(width - 15, height - 44, duration_string)
        canvas.drawRightString(width - 15, height - 58, f"Generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} +05:30")

        # Footer line and text
        canvas.line(15, 43, width - 15, 43)
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.black)
        canvas.drawString(15, 30, "ACME Inc")
        canvas.drawCentredString(width / 2, 30, f"Page {doc.page}")
        canvas.drawRightString(width - 15, 30, "Generated by Trisul Network Analytics (AI Edition)")

    # Shared table style
    base_table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2880BA")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ])

    elements = []
    for i, page in enumerate(pages):
        page_type = page.get("type", "")
        title = page.get("title", "")
        subtitle = page.get("subtitle", "")

        # Add titles
        elements.append(Spacer(1, 5))
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"<font color='#800080'>{subtitle}</font>", subtitle_style))
        elements.append(Spacer(1, 10))

        if page_type == "table":
            data = page.get("data", [])
            if not data:
                elements.append(Paragraph("<i>No table data available.</i>", styles["Normal"]))
            else:
                table = Table(
                    data,
                    repeatRows=1,
                    colWidths=[1.5*inch, 1.3*inch, 0.8*inch, 1.1*inch, 1.1*inch, 1.1*inch, 0.8*inch],
                )
                table.setStyle(base_table_style)
                elements.append(table)

        elif page_type == "chart":
            MAX_WIDTH = 6.5 * inch
            MAX_HEIGHT = 4.0 * inch
            

            image_path = page.get("file_path")
            if image_path:
                try:
                    # Read original image size
                    img_reader = ImageReader(image_path)
                    orig_w, orig_h = img_reader.getSize()

                    # Compute scale factor while preserving aspect ratio
                    scale_w = MAX_WIDTH / orig_w
                    scale_h = MAX_HEIGHT / orig_h
                    scale = min(scale_w, scale_h)

                    # Apply scaled dimensions
                    new_w = orig_w * scale
                    new_h = orig_h * scale

                    img = Image(image_path, width=new_w, height=new_h)
                    elements.append(img)
        
        
    
                except Exception as e:
                    elements.append(Paragraph(f"<i>Failed to load chart: {e}</i>", styles["Normal"]))
            else:
                elements.append(Paragraph("<i>No chart image path provided.</i>", styles["Normal"]))

        # Add page break except for the last page
        if i < len(pages) - 1:
            elements.append(PageBreak())

    pdf.build(elements, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)


    logging.info(f"[generate_trisul_report] PDF report generated at {filename}")

    return {"status": "success", "message" : f"The PDF report is generated successfully at {filename}. The report is displayed in the pop-up window, tell the user to kindly check that.", "file_path": filename}







@mcp.tool()
def search_keys(
    counter_group: str,
    pattern: Any = None,
    label: Any = None,
    keys: List[str] = None,
    maxitems: int = 100,
    offset: int = 0,
    get_totals: bool = False,
    get_attributes: bool = False,
    context: str = "context0",
    zmq_endpoint: str = None,
):
    """
    Search for a key in a counter group and its other values like description, label, and readable.
    We can provide a partial value or exact value to search for the key if the exact value is unknown.
    This tool returns a list of matching keys with their full details.

    How to use this tool for different search scenarios:
    
    1. Using the 'pattern' parameter (for Partial or Regex matches):
       - To match against 'key': Use a partial value. (e.g., pattern="42.23.3A"). NOTE: This is partial match only, no regex.
       - To match against 'label': Supports partial and regex match (e.g., pattern="^rtb.altitude").
       - To match against 'description': Supports partial and regex match (e.g., pattern="^test eedith*").

    2. Using the 'label' parameter (for Exact matches):
       - To match against 'key': Use the exact raw key (e.g., label="42.23.3A.53").
       - To match against 'label': Use the exact label (e.g., label="rtb.altitude-arena.com").

    Best Practices for better results:
    - Intelligent Counter Group Selection: Intelligently select the `counter_group` GUID based on what the user is asking for:
        * If 'link' is mentioned (e.g. 'bsnl link') → Use **FlowIntfs** (Flow Interface)
        * If 'host', 'user', or an IP is mentioned → Use **Hosts**
        * If it looks like an **ASN name** or number → Use **ASNumber**
        * If it looks like a **router** or **firewall** name → Use **Flowgens**
        * If it is a port or app name → Use **Apps**
        * **Clarification**: If you cannot determine the group even partially, DO NOT GUESS; ask the user to specify.
    - Case Sensitivity: Search can be case-sensitive. If searching for a name and it fails, try ALL CAPS or use a regex like '(?i)name' if supported.
    - Separators: If a name has spaces but might use hyphens (like 'GOOGLE-PRIVATE-CLOUD'), use '.*' in the pattern (e.g., pattern="(?i)google.*cloud").
    - Broad searching: If a specific name fails, search for the most unique single word in the name.
    - Ambiguous Matches: If multiple keys are returned (e.g., searching for interfaces across all routers), you MUST use a `get_counter_group_topper` query on that group (max 30 items, last 30 minutes) to identify which matching value has the highest recent activity. Prioritize and show data for that specific item first, then inform the user about other candidates.

    Parameters:
        counter_group (str): REQUIRED. GUID of the counter group to search in.
        pattern (str): Optional. Partial value or regex for matching key (partial only), label (regex), or description (regex).
        label (str): Optional. Exact value for matching key or human-readable label.
        keys (list[str]): Optional explicit list of raw key strings to look up directly.
        maxitems (int): Maximum number of keys to return (Default: 100).
        offset (int): Pagination offset — skip the first N results.
        get_totals (bool): If True, include aggregate metric totals for each key.
        get_attributes (bool): If True, include extended key attributes in the response.
        context (str): Trisul context identifier (Default: "context0").
        zmq_endpoint (str): Custom TRP ZMQ endpoint. Auto-computed if omitted.

    Returns:
        dict: SearchKeysResponse containing matching keys.
        Example result item structure:
            keys {
                key: "42.23.3A.53"
                readable: "66.35.58.83"
                label: "rtb.altitude-arena.com"
                description: "test eedith 123"
            }

    Example Scenarios:
        - High-recall name match: search_keys(cg, pattern="(?i).*microsoft.*corp.*")
        - Partial key match: search_keys(cg, pattern="42.23.3A", maxitems=20)
        - Regex label match: search_keys(cg, pattern="^rtb.altitude", maxitems=20)
        - Regex description match: search_keys(cg, pattern="^test eedith*", maxitems=20)
        - Exact key match: search_keys(cg, label="42.23.3A.53")
        - Exact label match: search_keys(cg, label="rtb.altitude-arena.com")
    """
    try:
        counter_group = str(counter_group)
        if pattern:
            pattern = str(pattern)
        if label:
            label = str(label)
        
        maxitems = int(maxitems)
        offset = int(offset)
        
        if not zmq_endpoint:
            ctx = normalize_context(context)
            zmq_endpoint = f"ipc:///usr/local/var/lib/trisul-hub/domain0/hub0/{ctx}/run/trp_0"

        logging.info(
            f"[search_keys] counter_group={counter_group} pattern={pattern} "
            f"label={label} keys={keys} maxitems={maxitems} offset={offset} "
            f"get_totals={get_totals} get_attributes={get_attributes} endpoint={zmq_endpoint}"
        )

        req = trp_pb2.Message()
        req.trp_command = req.SEARCH_KEYS_REQUEST
        q = req.search_keys_request

        q.counter_group = counter_group
        q.maxitems = int(maxitems)
        q.offset = int(offset)
        q.get_totals = bool(get_totals)
        q.get_attributes = bool(get_attributes)

        if pattern:
            q.pattern = pattern
            logging.info(f"[search_keys] pattern={pattern}")

        if label:
            q.label = label
            logging.info(f"[search_keys] label={label}")

        if keys:
            q.keys.extend([str(k) for k in keys])
            logging.info(f"[search_keys] explicit keys count={len(keys)}")

        logging.info("[search_keys] Sending SEARCH_KEYS_REQUEST")
        resp = get_response(zmq_endpoint, req)
        logging.info("[search_keys] Response received")
        
        # Post-process results for similarity if pattern was provided
        resp_dict = MessageToDict(resp)
        if pattern and 'keys' in resp_dict:
            import difflib
            search_str = pattern.lower()
            if pattern.startswith('(?i)'):
                search_str = pattern[4:].lower()
            
            scored_keys = []
            for k in resp_dict['keys']:
                label = k.get('label', '').lower()
                readable = k.get('readable', '').lower()
                key_id = k.get('key', '').lower()
                
                # Check label first, then readable, then key
                score = difflib.SequenceMatcher(None, search_str, label).ratio()
                score = max(score, difflib.SequenceMatcher(None, search_str, readable).ratio())
                score = max(score, difflib.SequenceMatcher(None, search_str, key_id).ratio())
                
                # Boost exact matches
                if search_str == label or search_str == readable or search_str == key_id:
                    score = 1.0
                
                k['similarity_score'] = round(score, 3)
                scored_keys.append((score, k))
            
            # Sort by score descending
            scored_keys.sort(key=lambda x: x[0], reverse=True)
            resp_dict['keys'] = [x[1] for x in scored_keys]
            
            # Add a hint for the LLM
            if resp_dict['keys'] and resp_dict['keys'][0]['similarity_score'] >= 0.5:
                resp_dict['suggestion'] = {
                    "best_match": resp_dict['keys'][0],
                    "confidence": "high" if resp_dict['keys'][0]['similarity_score'] >= 0.8 else "medium",
                    "note": f"Key '{resp_dict['keys'][0]['label']}' is a strong match for your search."
                }
        
        return json_to_toon(resp_dict)

    except Exception as e:
        logging.error(f"[search_keys] Error: {str(e)}", exc_info=True)
        return json_to_toon({"error": str(e)})


# AI Config tools

@mcp.tool()
def configure_llm_model():
    """
    Manage and switch between different LLM models for Trisul AI integrations.
    Usage:
        This tool allows administrators to switch the LLM model used in Trisul AI.
    
    Returns:
        str: Confirmation message.
    """
    logging.info(f"[configure_llm_model] Managing LLM model")
    return {"status": "success", "message" : f"The LLM model has been changed successfully."}

@mcp.tool()
def configure_embedding_model():
    """
    Manage and switch between different Embedding models for Trisul AI integrations.
    Usage:
        This tool allows administrators to switch the Embedding model used in Trisul AI.
    
    Returns:
        str: Confirmation message.
    """
    logging.info(f"[configure_embedding_model] Managing Embedding model")
    return {"status": "success", "message" : f"The Embedding model has been changed successfully."}

@mcp.tool()
def configure_llm_api_key():
    """
    Change the API key for the current LLM provider.
    Usage:
        This tool allows administrators to update the API key for the active LLM provider.

    Returns:
        str: Confirmation message.
    """
    logging.info(f"[configure_llm_api_key] Changing LLM API key")
    return {"status": "success", "message" : f"The LLM API key has been changed successfully."}

@mcp.tool()
def configure_embedding_api_key():
    """
    Change the API key for the current Embedding provider.
    Usage:
        This tool allows administrators to update the API key for the active Embedding provider.

    Returns:
        str: Confirmation message.
    """
    logging.info(f"[configure_embedding_api_key] Changing Embedding API key")
    return {"status": "success", "message" : f"The Embedding API key has been changed successfully."}


@mcp.tool()
def get_current_model_status():
    """
    Get the current model status for Trisul AI integrations.
    Usage:
        This tool allows administrators to get the current LLM and Embedding model status.
    Returns:
        str: Current model status.
    """
    logging.info(f"[get_current_model_status] Getting current model status")
    return {"status": "success", "message" : f"The current model status is --- ."}





if __name__ == "__main__":
    mcp.run(transport="stdio")
